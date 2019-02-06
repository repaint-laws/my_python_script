#!/usr/bin/env python
# coding: utf-8

import openpyxl
from random import shuffle

wb = openpyxl.Workbook()
ws = wb.active

list = [[a,b] for a in range(1,10) for b in range(1,10)]
shuffle(list)

fomulae = []
for v in list:
    fomulae.append("{0}+{1}=".format(v[0],v[1]))
for c in range(20):
    ws.append(fomulae[4*c:4*c+4])
# これで4列20行の計算式が完成

## 体裁を整える
# 空白列と行を挿入する
for i in range(1,4):
    ws.insert_cols(2*i)
for i in range(2):
    ws.insert_rows(11)

# セルを結合する
ws.merge_cells('D11:F11')
ws.merge_cells('D23:F23')
ws['D11'] = "分        秒"
ws['D23'] = "分        秒"

# サイズを大きくする
cols = ["A","B","C","D","E","F","G"]
ft = openpyxl.styles.fonts.Font(size=27)
for col in cols:
    for row in range(1,24):
        one_of_all = ws['{0}{1}'.format(col, row)]
        one_of_all.font = ft

wb.save("addition_test.xlsx")
